#!/usr/bin/env python
# coding: utf-8

# In[8]:


#!/usr/bin/env python
# coding: utf-8

# In[22]:


import tkinter as tk
from tkinter import filedialog
import openpyxl as op
import webbrowser
import ntpath
from openpyxl import Workbook
import os.path as os
from win32com import client
from pathlib import Path


root = tk.Tk()
root.title("엑셀파일 PDF로 변환하기")
root.geometry('580x400+500+400')

lab00=tk.Label(root,text="- 엑셀파일 PDF로 변환하기 -",font=('Arial 20 bold'),fg="black",width=34, height=2)
lab00.pack(side="top",pady=0)

def open():
    global my_excel
    root.filename = tk.filedialog.askopenfilename(initialdir="C:/Users/Public", title = "open file", filetypes = (("*.xlsx","*xlsx"),("*.xls","*xls"),("*.csv","*csv")))
 
    
    wb=op.load_workbook(root.filename)
    ws_names=wb.sheetnames   
    floce=os.split(root.filename)
    info=ws_names
    onlyname=Path(root.filename).stem
    counting=len(ws_names)
    flocation=floce[0]
    
    forname=onlyname
    filename=onlyname.replace(' ','_')
  

button01= tk.Button(root, text='파일열기', command=open,font=("Arial 16 bold"),width=40,height=2,fg="black",bg="yellow")
button01.pack(side="top",padx=1,pady=0)



def sheetsepertion():
    wb=op.load_workbook(root.filename)
    ws_names=wb.sheetnames   
    floce=os.split(root.filename)
    info=ws_names
    onlyname=Path(root.filename).stem
    counting=len(ws_names)
    flocation=floce[0]
    forname=onlyname
    filename=onlyname.replace(' ','_')
    
    print(info)
    sheetname=[]
    for j in info:
        data=j.replace(' ','_')
        sheetname.append(data)
        
    print(sheetname.append(data),sheetname)
    
    excel = client.Dispatch("Excel.Application")
    excel.Visible=False
    i=0
    
    for i in range(counting):
     
        sheets = excel.Workbooks.Open(root.filename) #파일 읽어오고
        work_sheets = sheets.Worksheets[i] 
        work_sheets.ExportAsFixedFormat(0, flocation+"/"+sheetname[i]+"_"+filename+".pdf")
        i +=1
        excel.Quit()
    
        


button04=tk.Button(root,text="PDF로 출력하기", font=("Arial 32 bold"),width=20,height=1,command=sheetsepertion,fg="white",bg="red")
button04.pack(side="top",pady=10)


def folderopen():
    wb=op.load_workbook(root.filename)
    ws_names=wb.sheetnames   
    floce=os.split(root.filename)
    info=ws_names
    onlyname=Path(root.filename).stem
    counting=len(ws_names)
    flocation=floce[0]
    path = flocation
    webbrowser.open(path)

def restart():
    filename="C:/Users/Public"

    

        





button09=tk.Button(root,text="PDF 파일 확인하기", font=("Arial 16 bold"),width=40,height=1,command=lambda:[folderopen(),restart()],fg="black",bg="orange")
button09.pack(side="top",pady=0)






lab05=tk.Label(root,text="제작 및 배포 : 언팩테크놀로지",font=('Arial 12 bold'),fg="blue",width=34)
lab05.pack(side="top",pady=10)

#버튼클릭시 링크오픈
def urllinked():
    print(webbrowser.open('https://unpac.kr', new=2))

button06=tk.Button(root,text="사용법 알아보기", font=("Arial 14 bold"),width=20,height=1,command=urllinked,fg="black",bg="lightblue")
button06.pack(side="top",pady=0)



root.mainloop()




# In[ ]:




